from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import datetime as dt
import os

fillable_names = ['CULTIVO', 'TAMAÑO', 'LOSA', 'GOLPE']
fillable_units = ['cubos', 'cm2', 'semillas/cubo']
individual_names = ['PROGENIE', 'FERTILIDAD', 'INVERSION', 'RENDIMIENTO', 'HABITAT']
collective_names = ['SIEMBRA', 'PROGENIE', 'FERTILIDAD', 'INVERSION', 'AREA', 'DENSIDAD', 'OCUPACION', 'COLONIZACION', 'RENDIMIENTO', 'HABITAT']
collective_units = ['semillas', 'plantas', 'plantas/semilla', 'semillas/planta', 'cm2', 'semillas/cm2', 'cm2/semilla', 'plantas/cubo', 'plantas/cm2', 'cm2/planta']

def create_template(parameters):

	wb = Workbook()
	ws = wb.active
	date = dt.datetime.now().strftime('%Y-%m-%d')
	ws.title = date

	ws['B1'] = 'SEMILLERO'

	crop, nrow, ncol = parameters

	ws.merge_cells('B1:' + chr(ord('B')+ncol) + '1')

	for col in range(3,3+ncol):
		ws[get_column_letter(col) + str(2)] = chr(ord('A')+(col-3))

	for row in range(3,3+nrow):
		ws['B' + str(row)] = str(row-2)

	for row in range(2,6):
		ws[get_column_letter(4+ncol) + str(row)] = fillable_names[row-2]

	for row in range(3,6):
		ws[get_column_letter(6+ncol) + str(row)] = fillable_units[row-3]

	for row in range(2,12):
		ws[get_column_letter(8+ncol) + str(row)] = collective_names[row-2]

	for row in range(2,12):
		ws[get_column_letter(10+ncol) + str(row)] = collective_units[row-2]

	ws[get_column_letter(5+ncol) + '2'] = crop
	ws[get_column_letter(5+ncol) + '3'] = nrow * ncol

	wb.save(f'../dat/templates/{crop}_{date}_{nrow}x{ncol}.xlsx')

def process_template(filename):
	
	wb = load_workbook('../dat/filled/' + filename)
	ws = wb.active

	name = filename.split('_')[0]
	date = filename.split('_')[1]
	nrow, ncol = list(map(int,filename.split('.')[0].split('_')[2].split('x')))
	size = nrow * ncol

	# read data
	individual = {}
	aux = []
	for row in range(3,3+nrow):
		for col in range(3,3+ncol):
			letter = get_column_letter(col)
			aux.append(ws[letter + str(row)].value)
	individual['PROGENIE'] = np.array(aux)

	fillable_data = []
	for row in range(2,6):
		fillable_data.append(ws[chr(ord('A')+(4+ncol)) + str(row)].value)

	fillable = dict(zip(fillable_names,fillable_data))

	# calculation of metrics
	individual['FERTILIDAD'] = individual['PROGENIE'] / fillable['GOLPE']
	individual['INVERSION'] = np.reciprocal(individual['FERTILIDAD'])
	individual['RENDIMIENTO'] = individual['PROGENIE'] / fillable['LOSA']
	individual['HABITAT'] = np.reciprocal(individual['RENDIMIENTO'])

	collective = {}
	collective['SIEMBRA'] = fillable['GOLPE']*fillable['TAMAÑO']
	collective['PROGENIE'] = np.sum(individual['PROGENIE'])
	collective['FERTILIDAD'] = collective['PROGENIE'] / collective['SIEMBRA']
	collective['INVERSION'] =  collective['SIEMBRA'] / collective['PROGENIE']
	collective['AREA'] = fillable['TAMAÑO'] * fillable['LOSA']
	collective['DENSIDAD'] = collective['SIEMBRA'] / collective['AREA']
	collective['OCUPACION'] = collective['AREA'] / collective['SIEMBRA']
	collective['COLONIZACION'] = collective['PROGENIE'] / fillable['TAMAÑO']
	collective['RENDIMIENTO'] = collective['PROGENIE'] / collective['AREA']
	collective['HABITAT'] = collective['AREA'] / collective['PROGENIE']

	# write collective data into excel sheet
	for row,value in zip(range(2,12),collective.values()):
		ws[chr(ord('A')+(8+ncol)) + str(row)] = value

	# write individual data into database
	with open('../dat/database.csv', 'a') as file:
		for i in range(size):
			file.write(name)
			file.write(',')
			file.write(date)
			file.write(',')
			for j in range(len(individual.keys())):
				file.write(str(np.around(individual[list(individual.keys())[j]][i], 2)))
				if j != len(individual.keys())-1:
					file.write(',')
			file.write('\n')

	wb.save('../dat/filled/' + filename)

def update_database(path):

	with open(path + 'database.csv', 'w') as file:

		file.write('CULTIVO')
		file.write(',')
		file.write('FECHA')
		file.write(',')

		for i in range(len(individual_names)):
			file.write(individual_names[i])
			if i != len(individual_names)-1:
				file.write(',')

		file.write('\n')

	files = os.listdir(path + 'templates/')

	for file in files:
		os.rename(path + 'templates/' + file, path + 'filled/' + file)

	files = os.listdir(path + 'filled/')

	for file in files:
		process_template(file)


#create_template(('poto',3,3))

update_database('../dat/')

df = pd.read_csv('../dat/database.csv', sep=',')
'''
df_cultivos = df.groupby(['CULTIVO'])
print(df_cultivos.get_group('poto'))
'''
sns.set_theme()
sns.set_style('ticks')
sns.set_context('paper')
sns.boxplot(data=df, x='CULTIVO', y='FERTILIDAD')
sns.stripplot(data=df, x='CULTIVO', y='FERTILIDAD', color='black')
sns.despine(top=True, right=True, offset=0, trim=False)

plt.show()
