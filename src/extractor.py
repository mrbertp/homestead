from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import numpy as np

filename = 'sample_entry.xlsx'

wb = load_workbook(filename)
sheets = wb.sheetnames

for sheet in sheets:
	ws = wb[sheet]
	#ws.title = ws.title + ' ' + ws['O2'].value

	data = []

	for row in range(4,14):
		for col in range(4,14):
			letter = get_column_letter(col)
			data.append(ws[letter + str(row)].value)

	data = np.array(data)

	stats = {'Media': round(np.mean(data),2),
			 'Mediana': round(np.median(data),2),
			 'Des. Est.': round(np.std(data),2),
			 'P': np.sum(data),
			 'F': np.sum(data)/(data.size*ws['P4'].value)
			 }

	ws['P6'] = stats['Media']
	ws['P7'] = stats['Mediana']
	ws['P8'] = stats['Des. Est.']
	ws['P9'] = stats['P']
	ws['P10'] = stats['F']

wb.save(filename)